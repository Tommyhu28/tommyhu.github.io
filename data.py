import datetime
from openpyxl import load_workbook

path='static/member.xlsx'
workbook = load_workbook(filename=path)
sheet = workbook.active
member=workbook.get_sheet_by_name('Member')
post=workbook.get_sheet_by_name('Posts')

member_list=[]
for row in member.iter_rows(min_row=2,max_row=6,max_col=5,values_only=True):
  member_list.append(row)

post_list=[]
for row in post.iter_rows(min_row=2,values_only=True):
  post_list.append(row)

ten_posts=[]
for row in post.iter_rows(min_row=2,max_row=11,values_only=True):
  ten_posts.append(row)

# print(member_list[1])
# print(post_list[0])
# print(post_list)

def sort_find(collection):
  list=[]
  sorted_posts=[]
  for section in collection:
    list.append(section[5]) # hard-coded index for date
  list.sort(key=lambda date: datetime.datetime.strptime(date, '%m %d %Y'),reverse=True)
  for item in list:
    for part in collection:
      if part[5]==item:
        sorted_posts.append(part)
  return sorted_posts


def organize(collection):
  final=[]
  for section in post_list:
    temp=[]
    temp.append(section[0])
    temp.append(section[1])
    temp.append(section[2].split('|'))
    temp.append(section[3].split('|'))
    temp.append(section[4].split('|'))
    temp.append(section[5])
    final.append(temp)

  return sort_find(final)

organized_posts=organize(post_list)
organized_ten=organize(ten_posts)

# to obtain a particular user's post info
def find_post(name):
  result=[]
  for section in organized_posts:
    if section[0]==name: # hard-coded
      result.append(section)

  return result

# successfully extracted and categorized each part of a post into sections below:
# rtr=find_post('Cloud')
# print(rtr)

def get_members():
  list=[]
  for section in member_list:
    list.append(section)
  return list

# for displaying date
def date_format(value, format='medium'):
  date_obj = datetime.datetime.strptime(value, '%m %d %Y')
  if format == 'full':
    return date_obj.strftime("%m/%d/%Y")
  elif format == 'medium':
    return date_obj.strftime("%b %d %Y")      

# uses a name to return a cat's info
def get_cat(input):
  for section in member_list:
    if section[0]==input:
      return section
  return None
